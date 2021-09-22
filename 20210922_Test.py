import argparse
import os
import glob
import random
import darknet
import time
import cv2
import numpy as np
import darknet
import spidev
import sys
import serial
import openpyxl
from openpyxl import Workbook
from openpyxl import load_workbook
import requests
from datetime import datetime
import json
import smbus
import Jetson.GPIO as GPIO
url_Weight = "http://ibp.bime.ntu.edu.tw/rest/sensorDataLogs/CH01/PH01/weight_test"
url_image = "http://ibp.bime.ntu.edu.tw/rest/sensorDataLogs/CH01/PH01/weight_test/image/file"
url_AVG_Weight = "http://ibp.bime.ntu.edu.tw/rest/sensorDataLogs/CH01/PH01/weight_test/average_weight"
url_image_NoFlash = "http://ibp.bime.ntu.edu.tw/rest/sensorDataLogs/CH01/PH01/weight_test/noflash_image/file"
headers = {"Content-Type":"application/json"}
LED_COUNT = 24
d1=time.strftime("%Y-%m-%d %H:%M:%S")


### Upload_img_to_IBP
def upload_img_to_IBP(dataName,image ,url):    
  #dataTime = time.strftime("%Y-%m-%d %H:%M:%S")
  query = {'dataTime' : d1}
  files = {'file': open(dataName, 'rb')}
  try:
    res = requests.post(url, files=files, data=query, timeout=20)
    print('Image Upload Successed')
  except:
    print('Image Upload Failed')  

### Upload_data_to_IBP
def upload_data_to_IBP(dataName, data, url):     
  dataTime = time.strftime("%Y-%m-%d %H:%M:%S")
  query = {'dataTime':d1, dataName:"{:.2f}".format(data)}
  query_j = json.dumps(query)
  print(query_j)
  try:
    res = requests.post(url, data=query_j,headers=headers)
    print('Data Upload Successed')
  except:
    print('Data Upload Failed')

### LIGHT
class SPItoWS():
    def __init__(self, ledc):

        self.led_count = ledc

        self.X = ''  # X is signal of WS281x
        for i in range(self.led_count):
            self.X = self.X + "100100100100100100100100100100100100100100100100100100100100100100100100"
        self.spi = spidev.SpiDev()
        self.spi.open(1, 1)  # SPI2---->MOSI:37pin (/dev/SPI1.1)
        self.spi.max_speed_hz = 2400000

    def __init__(self, ledc , busIn=1, deviceIn=1):

        self.led_count = ledc
        self.bus = busIn
        self.device = deviceIn
        self.X = '' # X is signal of WS281x
        for i in range(self.led_count):
            self.X = self.X + "100100100100100100100100100100100100100100100100100100100100100100100100"
        self.spi = spidev.SpiDev()
        self.spi.open(self.bus, self.device)     #SPI2---->MOSI:37pin (/dev/SPI1.1)
        self.spi.max_speed_hz = 2400000

    def __del__(self):
        self.spi.close()

    def _Bytesto3Bytes(self, num, RGB): # num is number of signal, RGB is 8 bits (1 byte) str
        for i in range(8):
            if RGB[i] == '0':
                self.X = self.X[:num * 3 * 8 + i * 3] + '100' + self.X[num * 3 * 8 + i * 3 + 3:]
            elif RGB[i] == '1':
                self.X = self.X[:num * 3 * 8 + i * 3] + '110' + self.X[num * 3 * 8 + i * 3 + 3:]

    def _BytesToHex(self, Bytes):
        return ''.join(["0x%02X " % x for x in Bytes]).strip()

    def LED_show(self):
            Y = []
            for i in range(self.led_count * 9):
                Y.append(int(self.X[i*8:(i+1)*8],2))
            WS = self._BytesToHex(Y)
            self.spi.xfer3(Y, 2400000,0,8)

    def LED_show(self, R, G, B):
        for LED_NUM in range(LED_COUNT):
            self.RGBto3Bytes(LED_NUM, 255, 255, 255)
            self.LED_show()
            time.sleep(.1)

    def RGBto3Bytes(self, led_num, R, G, B):
        if (R > 255 or G > 255 or B > 255):
            print("Invalid Value: RGB is over 255\n")
            sys.exit(1)
        if (led_num > self.led_count - 1):
            print("Invalid Value: The number is over the number of LED")
            sys.exit(1)
        RR = format(R, '08b')
        GG = format(G, '08b')
        BB = format(B, '08b')
        self._Bytesto3Bytes(led_num * 3, GG)
        self._Bytesto3Bytes(led_num * 3 + 1, RR)
        self._Bytesto3Bytes(led_num * 3 + 2, BB)

    def LED_OFF_ALL(self):
        self.X = ''
        for i in range(self.led_count):
            self.X = self.X + "100100100100100100100100100100100100100100100100100100100100100100100100"
        self.LED_show()



### Camera parameter
def gstreamer_pipeline(
    capture_width=3264,
    capture_height=2040,
    display_width=3264,
    display_height=2040,
    framerate=5,
    flip_method=1,
):
    return (
        "nvarguscamerasrc ! "
        "video/x-raw(memory:NVMM), "
        "width=(int)%d, height=(int)%d, "
        "format=(string)NV12, framerate=(fraction)%d/1 ! "
        "nvvidconv flip-method=%d ! "
        "video/x-raw, width=(int)%d, height=(int)%d, format=(string)BGRx ! "
        "videoconvert ! "
        "video/x-raw, format=(string)BGR ! appsink"
        % (
            capture_width,
            capture_height,
            framerate,
            flip_method,
            display_width,
            display_height,
        )
    )

### Take a picture and Upload
def show_camera():
    cap = cv2.VideoCapture(gstreamer_pipeline(flip_method=0), cv2.CAP_GSTREAMER)
    ret_val, img = cap.read()
    images_file='/home/ichase/original_picture'
    images_file1='/home/ichase/temporary_picture'
    cv2.imwrite(images_file+'/'+d1+'.jpg',img)
    cv2.imwrite(images_file1+'/'+d1+'.jpg',img)
   
def show_camera1():
    cap = cv2.VideoCapture(gstreamer_pipeline(flip_method=0), cv2.CAP_GSTREAMER)
    ret_val, img = cap.read()
    
    images_file3='/home/ichase/noflash_picture'
    
    cv2.imwrite(images_file3+'/'+d1+'.jpg',img)
    
        
    #upload_img_to_IBP(img_dataname,url1)

### Machine Learning
def parser():
    parser = argparse.ArgumentParser(description="YOLO Object Detection")
    parser.add_argument("--input", type=str, default="",
                        help="image source. It can be a single image, a"
                        "txt with paths to them, or a folder. Image valid"
                        " formats are jpg, jpeg or png."
                        "If no input is given, ")
    parser.add_argument("--batch_size", default=1, type=int,
                        help="number of images to be processed at the same time")
    parser.add_argument("--weights", default="yolov4-obj-hen_best.weights",
                        help="yolo weights path")
    parser.add_argument("--dont_show", action='store_true',
                        help="windown inference display. For headless systems")
    parser.add_argument("--ext_output", action='store_true',
                        help="display bbox coordinates of detected objects")
    parser.add_argument("--save_labels", action='store_true',
                        help="save detections bbox for each image in yolo format")
    parser.add_argument("--config_file", default="./cfg/yolov4.cfg",
                        help="path to config file")
    parser.add_argument("--data_file", default="./cfg/coco.data",
                        help="path to data file")
    parser.add_argument("--thresh", type=float, default=.25,
                        help="remove detections with lower confidence")
    return parser.parse_args()


def check_arguments_errors(args):
    assert 0 < args.thresh < 1, "Threshold should be a float between zero and one (non-inclusive)"
    if not os.path.exists(args.config_file):
        raise(ValueError("Invalid config path {}".format(os.path.abspath(args.config_file))))
    if not os.path.exists(args.weights):
        raise(ValueError("Invalid weight path {}".format(os.path.abspath(args.weights))))
    if not os.path.exists(args.data_file):
        raise(ValueError("Invalid data file path {}".format(os.path.abspath(args.data_file))))
    if args.input and not os.path.exists(args.input):
        raise(ValueError("Invalid image path {}".format(os.path.abspath(args.input))))


def check_batch_shape(images, batch_size):
    """
        Image sizes should be the same width and height
    """
    shapes = [image.shape for image in images]
    if len(set(shapes)) > 1:
        raise ValueError("Images don't have same shape")
    if len(shapes) > batch_size:
        raise ValueError("Batch size higher than number of images")
    return shapes[0]


def load_images(images_path):
    """
    If image path is given, return it directly
    For txt file, read it and return each line as image path
    In other case, it's a folder, return a list with names of each
    jpg, jpeg and png file
    """
    input_path_extension = images_path.split('.')[-1]
    if input_path_extension in ['jpg', 'jpeg', 'png']:
        return [images_path]
    elif input_path_extension == "txt":
        with open(images_path, "r") as f:
            return f.read().splitlines()
    else:
        return glob.glob(
            os.path.join(images_path, "*.jpg")) + \
            glob.glob(os.path.join(images_path, "*.png")) + \
            glob.glob(os.path.join(images_path, "*.jpeg"))


def prepare_batch(images, network, channels=3):
    width = darknet.network_width(network)
    height = darknet.network_height(network)

    darknet_images = []
    for image in images:
        image_rgb = cv2.cvtColor(image, cv2.COLOR_BGR2RGB)
        image_resized = cv2.resize(image_rgb, (width, height),
                                   interpolation=cv2.INTER_LINEAR)
        custom_image = image_resized.transpose(2, 0, 1)
        darknet_images.append(custom_image)

    batch_array = np.concatenate(darknet_images, axis=0)
    batch_array = np.ascontiguousarray(batch_array.flat, dtype=np.float32)/255.0
    darknet_images = batch_array.ctypes.data_as(darknet.POINTER(darknet.c_float))
    return darknet.IMAGE(width, height, channels, darknet_images)


def image_detection(image_path, network, class_names, class_colors, thresh):
    # Darknet doesn't accept numpy images.
    # Create one with image we reuse for each detect
    
    width = darknet.network_width(network)
    height = darknet.network_height(network)
    darknet_image = darknet.make_image(width, height, 3)

    image = cv2.imread(image_path)
    image_rgb = cv2.cvtColor(image, cv2.COLOR_BGR2RGB)
    image_resized = cv2.resize(image_rgb, (width, height),
                               interpolation=cv2.INTER_LINEAR)

    darknet.copy_image_from_bytes(darknet_image, image_resized.tobytes())
    detections = darknet.detect_image(network, class_names, darknet_image, thresh=thresh)
    darknet.free_image(darknet_image)
    
    image = darknet.draw_boxes(detections, image_resized, class_colors)
        
    return cv2.cvtColor(image, cv2.COLOR_BGR2RGB), detections


def batch_detection(network, images, class_names, class_colors,
                    thresh=0.9, hier_thresh=.5, nms=.45, batch_size=4):
    image_height, image_width, _ = check_batch_shape(images, batch_size)
    darknet_images = prepare_batch(images, network)
    batch_detections = darknet.network_predict_batch(network, darknet_images, batch_size, image_width,
                                                     image_height, thresh, hier_thresh, None, 0, 0)
    batch_predictions = []
    for idx in range(batch_size):
        num = batch_detections[idx].num
        detections = batch_detections[idx].dets
        if nms:
            darknet.do_nms_obj(detections, num, len(class_names), nms)
        predictions = darknet.remove_negatives(detections, class_names, num)
        images[idx] = darknet.draw_boxes(predictions, images[idx], class_colors)
        batch_predictions.append(predictions)
    darknet.free_batch_detections(batch_detections, batch_size)
    return images, batch_predictions


def image_classification(image, network, class_names):
    
    width = darknet.network_width(network)
    height = darknet.network_height(network)
    image_rgb = cv2.cvtColor(image, cv2.COLOR_BGR2RGB)
    image_resized = cv2.resize(image_rgb, (width, height),
                                interpolation=cv2.INTER_LINEAR)
    darknet_image = darknet.make_image(width, height, 3)
    darknet.copy_image_from_bytes(darknet_image, image_resized.tobytes())
    detections = darknet.predict_image(network, darknet_image)
    predictions = [(name, detections[idx]) for idx, name in enumerate(class_names)]
    darknet.free_image(darknet_image)

   
    return sorted(predictions, key=lambda x: -x[1])


def convert2relative(image, bbox):
    """
    YOLO format use relative coordinates for annotation
    """
    x, y, w, h = bbox
    height, width, _ = image.shape
    return x/width, y/height, w/width, h/height


def save_annotations(name, image, detections, class_names):
    """
    Files saved with image_name.txt and relative coordinates
    """
    file_name = os.path.splitext(name)[0] + ".txt"
    with open(file_name, "w") as f:
        for label, confidence, bbox in detections:
            x, y, w, h = convert2relative(image, bbox)
            label = class_names.index(label)
            f.write("{} {:.4f} {:.4f} {:.4f} {:.4f} {:.4f}\n".format(label, x, y, w, h, float(confidence)))


def batch_detection_example():
    args = parser()
    check_arguments_errors(args)
    batch_size = 3
    random.seed(3)  # deterministic bbox colors
    network, class_names, class_colors = darknet.load_network(
        args.config_file,
        args.data_file,
        args.weights,
        batch_size=batch_size
    )
    image_names = ['data/horses.jpg', 'data/horses.jpg', 'data/eagle.jpg']
    images = [cv2.imread(image) for image in image_names]
    images, detections,  = batch_detection(network, images, class_names,
                                           class_colors, batch_size=batch_size)
    for name, image in zip(image_names, images):
        cv2.imwrite(name.replace("data/", ""), image)
    print(detections)


### hx711
class JoeI2C:

    def __init__(self, address = 0x08,  busNum=0, ackPin=29):
        # Since the I2C protocol in Jetson nano just performs master mode, a ack pin is needed to tell arduino to read
        # something sent from jetson nano
        self.address = address
        self.busNum = busNum
        self.ackPin = ackPin
        # Nvidia Jetson Nano i2c Bus 0 (default)
        self.bus = smbus.SMBus(busNum)

        ##========================================== Definition ================================#
        # Register number define
        self.REG_CURMODE            =   0x00    # (W/R)
        self.REG_JN_CORE_TEMP       =   0x01    # (W)
        self.REG_AVG_WEIGHT         =   0x02    # (W/R)
        self.REG_CUR_TOTAL_WEIGHT   =   0x03    # (R)
        self.REG_OFFSET             =   0x04    # (R)
        self.REG_CUR_JN_STAT        =   0xFE    # (W)
        self.REG_SHUTDOWN_CMD       =   0xFD    # (R) Read this Reg will trigger Atmega send Shutdown command

        # Mode define
        self.MODE_NONE          =   0x00    # No mode is selected
        self.MODE_MEASURE       =   0x01    # IN measuring mode
        self.MODE_SLEEP         =   0x02
        self.MODE_SHUTDOWN      =   0xFD

        # Jetson Nano Status
        self.JNSTAT_WORKWELL    =   0x00
        self.JNSTAT_RESET       =   0x01    # Jetson Nano has been reseted once
        self.JNSTAT_CAMERAFAIL  =   0x02
        self.JNSTAT_WIFIFAIL    =   0x03
        # default address of Arduino is 0x08
        GPIO.setwarnings(False)
        GPIO.setmode(GPIO.BOARD)  # BOARD pin-numbering scheme
        GPIO.setup(self.ackPin,GPIO.IN)
        GPIO.add_event_detect(self.ackPin, GPIO.FALLING, callback=self.I2C_Event)

        # Send msg to change the mode to MODE_MEASURE
        data = [0x00, self.MODE_MEASURE]
        self.bus.write_i2c_block_data(address, self.REG_CURMODE, data)


    def I2C_Event(self, channel):
        GPIO.remove_event_detect(self.ackPin)
        print("I2C Int!!")
        command = bytes(self.readNumber(self.REG_SHUTDOWN_CMD, 2))
        if command[0] == 0x15:
            if command[1] == 0x65:
                print("shutdown!!")
                time.sleep(3)
                os.system('shutdown -h now')

        print("0x" + command.hex())
        time.sleep(0.001)
        GPIO.add_event_detect(self.ackPin, GPIO.FALLING, callback=self.I2C_Event)

    def writeNumber(self,Reg, value):
        #bus.write_byte(address, value)
        self.bus.write_byte_data(self.address, 0, value)
        return -1

    def readNumber(self, Cmd,NumBytes):
        number = self.bus.read_i2c_block_data(self.address, Cmd, NumBytes)
        # number = bus.read_byte_data(address, 1)
        return number

    def readWeight(self):
        Value = np.uint16(self.readNumber(self.REG_CUR_TOTAL_WEIGHT, 2))
        #print ("Value: " + "0x{:02X} ".format(Value[0]) + "0x{:02X} ".format(Value[1]))
        Temp = np.int16(((Value[0]<<8) & 0xFF00) | (Value[1] & 0x00FF))
        #print("Temp: " + str(Temp))
        Weight = float(Temp/100)
        #print("Weight :" + str(Weight))
        return Weight

    def writeAvgWeight(self, AvgWeight):
        #print("AvgWeight : " + str(AvgWeight))
        Data_toSend = np.uint16(AvgWeight*pow(10,2))
        #print("Data_toSend : " + str(Data_toSend))
        byte_High = int((Data_toSend>>8) & 0xFF)
        byte_Low = int((Data_toSend) & 0xFF)
        data = [byte_High, byte_Low]
        self.bus.write_i2c_block_data(self.address, self.REG_AVG_WEIGHT, data)
        print ("Value: " + "0x{:02X} ".format(byte_High) + "0x{:02X} ".format(byte_Low))
        
    def offset(self):
        success = self.readNumber(self.REG_OFFSET, 2)
        if success[0] == 0x13:
            if success[1] == 0x65:
                print("offset success!!")      


    

### get weight
def weight():
    try:
        while True:
            Weight = myI2C.readWeight()
            print("Weight :" + str(Weight))
            #time.sleep(5)
            
             #bus.write_byte_data(address, 0x15, 0x85)
            #number = readNumber(0x85,3)
            #print(number)
            return Weight
        # This is the address we setup in the Arduino Program


    except KeyboardInterrupt:
        print("Exiting Program")

    # except Exception as exception_error:
    #     print("Error occurred. Exiting Program")
    #     print("Error: " + str(exception_error))

    finally:
        # MySerial.Close()
        pass

    

def main():
    try:   
        #Defining the path of
        images_file='/home/ichase/original_picture' ###original picture
        images_file1='/home/ichase/temporary_picture' ### temporary picture
        images_file2='/home/ichase/final_picture' ### final picture
        images_file3='/home/ichase/noflash_picture'

        #Initialize the WS2812B LED Strip
        sig = SPItoWS(LED_COUNT)
        
        sig.LED_OFF_ALL()   
        sig.LED_show(255,255,255)
        time.sleep(1)
        #take a picture and upload
        result = weight()
        show_camera()
        #get the weight
        sig.LED_OFF_ALL()
        time.sleep(1)
        result_NoFalsh = weight()
        show_camera1()
        #get the weight
        
        
        #upload the weight and time to IBP
        upload_data_to_IBP('weight_test', result, url_Data)

        
        
        args = parser()
        check_arguments_errors(args)

        random.seed(3)  # deterministic bbox colors
        network, class_names, class_colors = darknet.load_network(
            args.config_file,
            args.data_file,
            args.weights,
            batch_size=args.batch_size
        )

        images = load_images(args.input)
    
        index = 0
        
        while True:
            ### loop asking for new image paths if no list is given
            if args.input:
                if index >= len(images):
                    break
                image_name = images[index]
            else:
                image_name = (images_file1 + '/'+d1+'.jpg')
                print(image_name)
            prev_time = time.time()    
            
        ### Using cv2.putText() method
            
            image, detections = image_detection( image_name, network, class_names, class_colors, args.thresh  )
            count = 0
            #x, y, w, h = convert2relative(image, bbox)
            for label, confidence, bbox in detections:
                x, y, w, h = convert2relative(image, bbox)
                print(x,y,w,h)     
                if 140<x*512<380:
                    if 80<y*288<250: 
                        count+=1
            print("+++++", count)

            ### find the excel to record data
            wb = load_workbook('filename'+'.xlsx')
            ws = wb.active
            #ws = wb.get_sheet_by_name('Sheet')
            k = ws.max_row
            ws.cell(row=k+1,column=1).value=d1 ### date
            ws.cell(row=k+1,column=2).value=result ### total weight
            ws.cell(row=k+1,column=3).value=count ### animal count
            average = 0
            if count == 0:
                ws.cell(row=k+1,column=4).value=0
                
            else:
                
                average =float(result/count)
                average_str = ('%.2f' %average)
                ws.cell(row=k+1,column=4).value=average_str ### average per one
                average = round(average, 2)
                print(average)
            myI2C.writeAvgWeight(average)
            upload_data_to_IBP('average_weight', average, url)
                
            ### close the excel and save       
            wb.close()   
            wb.save('filename'+'.xlsx') 

        

            if args.save_labels:
                save_annotations(image_name, image, detections, class_names)
            darknet.print_detections(detections, args.ext_output)
            fps = int(1/(time.time() - prev_time))
            print("FPS: {}".format(fps))
            print(d1)
            cv2.imwrite(images_file1+'/'+d1+'.jpg',image)
            #cv2.imshow('Inference', image)
            #if cv2.waitKey() & 0xFF == ord('q'):

            ### make a rectangle 
            ret=cv2.imread(images_file1+'/'+d1+'.jpg')
            cv2.rectangle(ret,(140,80),(380,250),(255,0,0),2)
            cv2.imwrite(images_file1+'/'+d1+'.jpg',ret)

            ### make the black rectange on the upper left corner and putText the word 
            pho = cv2.imread(images_file1+'/'+d1+'.jpg')
            width = pho.shape[1]
            height = pho.shape[2]
            for y in range (0,30):
                for x in range (0,250):
                    pho[y,x] = 0
            font = cv2.FONT_HERSHEY_PLAIN
            img1 = cv2.putText(pho,'weight='+str(result)+'count='+str(count)+ ' '+'average_weight='+str(average),(30,25),font,2,(255,255,255),1)
            cv2.imwrite(images_file1+'/'+d1+'.jpg',img1)
            
            pho = cv2.imread(images_file1+'/'+d1+'.jpg')
            img = cv2.putText(pho,d1,(30,10),font,2,(255,255,255),1)
            cv2.imwrite(images_file2+'/'+d1+'.jpg',img)
            
            ####
            
            pho = cv2.imread(images_file+'/'+d1+'.jpg')
            width = pho.shape[1]
            height = pho.shape[2]
            for y in range (0,100):
                for x in range (0,1200):
                    pho[y,x] = 0
            font = cv2.FONT_HERSHEY_PLAIN
            image = cv2.putText(pho,'Total weight='+str(result)+' '+'count='+str(count)+ ' '+'average_weight='+str(average),(30,75),font,3,(255,255,255),1)
            cv2.imwrite(images_file+'/'+d1+'.jpg',pho)
            
            pho = cv2.imread(images_file+'/'+d1+'.jpg')
            img = cv2.putText(pho,d1,(30,30),font,3,(255,255,255),1)
            cv2.imwrite(images_file+'/'+d1+'.jpg',img)
            dataName = images_file+'/'+d1+'.jpg'
            upload_img_to_IBP(dataName, url1)
            ###noflash
            
            pho = cv2.imread(images_file3+'/'+d1+'.jpg')
            width = pho.shape[1]
            height = pho.shape[2]
            for y in range (0,100):
                for x in range (0,1200):
                    pho[y,x] = 0
            font = cv2.FONT_HERSHEY_PLAIN
            image = cv2.putText(pho,'Total weight='+str(result)+' '+'count='+str(count)+ ' '+'average_weight='+str(average),(30,75),font,3,(255,255,255),1)
            cv2.imwrite(images_file3+'/'+d1+'.jpg',pho)
            
            pho = cv2.imread(images_file3+'/'+d1+'.jpg')
            img = cv2.putText(pho,d1,(30,30),font,3,(255,255,255),1)
            cv2.imwrite(images_file3+'/'+d1+'.jpg',img)

            dataName1 = images_file3+'/'+d1+'.jpg'
            upload_img_to_IBP(dataName1, url3)
            
            
            index += 1
            break

    except KeyboardInterrupt:
        print("Exiting Program")

if __name__ == "__main__":
    # unconmment next line for an example of batch processing
    # batch_detection_example()
    #show_camera()
    myI2C = JoeI2C()
    main()
    